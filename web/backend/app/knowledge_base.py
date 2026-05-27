"""
OpsBrain — Knowledge Base (SQLite FTS5)

使用 SQLite FTS5 全文搜索。零额外依赖，无需向量模型。
支持中英文混合搜索。

数据存储在知识库自有的独立 SQLite 文件中以防冲突。
"""

from __future__ import annotations

import csv
import io
import json
import os
import sqlite3
import re

KNOWLEDGE_DIR = os.environ.get("OPSBRAIN_HOME", "/var/lib/opsbrain") + "/knowledge"
_DB_PATH = None
_conn: sqlite3.Connection | None = None


def _ensure_dir():
    os.makedirs(KNOWLEDGE_DIR, exist_ok=True)


def _get_db() -> sqlite3.Connection:
    global _conn, _DB_PATH
    _ensure_dir()
    if _conn is None:
        _DB_PATH = KNOWLEDGE_DIR + "/knowledge.db"
        _conn = sqlite3.connect(_DB_PATH, check_same_thread=False)
        _conn.row_factory = sqlite3.Row
        _conn.execute("PRAGMA journal_mode=WAL")
        _init_schema()
    return _conn


def _init_schema():
    """初始化 FTS5 表"""
    conn = _get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS configs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor TEXT NOT NULL DEFAULT '*',
            task TEXT NOT NULL,
            commands TEXT NOT NULL,
            notes TEXT DEFAULT ''
        );
        CREATE VIRTUAL TABLE IF NOT EXISTS configs_fts USING fts5(
            vendor, task, commands, notes,
            content=configs, content_rowid=id
        );
        -- 自动同步触发器
        CREATE TRIGGER IF NOT EXISTS configs_ai AFTER INSERT ON configs BEGIN
            INSERT INTO configs_fts(rowid, vendor, task, commands, notes)
            VALUES (new.id, new.vendor, new.task, new.commands, new.notes);
        END;
        CREATE TRIGGER IF NOT EXISTS configs_ad AFTER DELETE ON configs BEGIN
            INSERT INTO configs_fts(configs_fts, rowid, vendor, task, commands, notes)
            VALUES ('delete', old.id, old.vendor, old.task, old.commands, old.notes);
        END;
        CREATE TRIGGER IF NOT EXISTS configs_au AFTER UPDATE ON configs BEGIN
            INSERT INTO configs_fts(configs_fts, rowid, vendor, task, commands, notes)
            VALUES ('delete', old.id, old.vendor, old.task, old.commands, old.notes);
            INSERT INTO configs_fts(rowid, vendor, task, commands, notes)
            VALUES (new.id, new.vendor, new.task, new.commands, new.notes);
        END;
    """)
    conn.commit()


# ═══ 预置配置模板 ══════════════════════════════════════════

_PRESET_CONFIGS = [
    {"vendor": "cisco", "task": "端口状态检查", "commands": "show interface status\nshow interface description\nshow ip interface brief", "notes": "检查所有端口状态和IP接口摘要"},
    {"vendor": "cisco", "task": "VLAN配置", "commands": "vlan {vlan_id}\nname {vlan_name}\ninterface {interface}\nswitchport mode access\nswitchport access vlan {vlan_id}\nno shutdown", "notes": "创建VLAN并分配端口。替换{vlan_id},{vlan_name},{interface}"},
    {"vendor": "cisco", "task": "Trunk配置", "commands": "interface {interface}\nswitchport mode trunk\nswitchport trunk allowed vlan {vlan_list}\nswitchport trunk native vlan {native_vlan}", "notes": "配置Trunk端口"},
    {"vendor": "cisco", "task": "路由表检查", "commands": "show ip route\nshow ip route summary\nshow ip protocols", "notes": "查看路由表和路由协议状态"},
    {"vendor": "cisco", "task": "生成树配置", "commands": "spanning-tree vlan {vlan_id} root primary\nspanning-tree portfast default", "notes": "配置STP根桥和PortFast"},
    {"vendor": "cisco", "task": "ACL配置", "commands": "ip access-list extended {acl_name}\npermit ip {src} {src_mask} {dst} {dst_mask}\ninterface {interface}\nip access-group {acl_name} in", "notes": "创建并应用扩展ACL"},
    {"vendor": "cisco", "task": "OSPF配置", "commands": "router ospf {pid}\nnetwork {network} {wildcard} area {area}\ninterface {interface}\nip ospf priority 0", "notes": "配置OSPF邻居和区域"},
    {"vendor": "cisco", "task": "端口安全", "commands": "switchport port-security\nswitchport port-security maximum 2\nswitchport port-security violation restrict\nswitchport port-security mac-address sticky", "notes": "端口安全配置，限制MAC数量"},
    {"vendor": "cisco", "task": "EtherChannel", "commands": "interface port-channel {num}\ninterface {interface}\nchannel-group {num} mode active", "notes": "LACP链路聚合"},
    {"vendor": "cisco", "task": "NAT配置", "commands": "ip nat inside source list {acl} interface {interface} overload\ninterface {inside}\nip nat inside\ninterface {outside}\nip nat outside", "notes": "PAT/NAT过载配置"},
    {"vendor": "cisco", "task": "BGP配置", "commands": "router bgp {asn}\nneighbor {ip} remote-as {remote_asn}\nnetwork {network} mask {mask}", "notes": "BGP路由协议配置"},
    {"vendor": "cisco", "task": "DHCP配置", "commands": "ip dhcp pool {pool_name}\nnetwork {network} {mask}\ndefault-router {gateway}\ndns-server {dns}\nlease {days}", "notes": "DHCP服务器配置"},
    {"vendor": "cisco", "task": "接口IP配置", "commands": "interface {interface}\nip address {ip} {mask}\nno shutdown", "notes": "配置接口IP地址"},
    {"vendor": "cisco", "task": "静态路由", "commands": "ip route {destination} {mask} {next_hop}", "notes": "配置静态路由"},
    {"vendor": "cisco", "task": "SSH配置", "commands": "hostname {name}\nip domain-name {domain}\ncrypto key generate rsa modulus 2048\nip ssh version 2\nline vty 0 4\ntransport input ssh\nlogin local", "notes": "开启SSH远程管理"},
    {"vendor": "cisco", "task": "SNMP配置", "commands": "snmp-server community {community} {ro|rw}\nsnmp-server location {location}\nsnmp-server contact {email}", "notes": "SNMP只读/读写配置"},
    {"vendor": "cisco", "task": "NTP配置", "commands": "ntp server {server_ip}\nntp update-calendar", "notes": "NTP时间同步"},
    {"vendor": "cisco", "task": "远程日志", "commands": "logging host {server_ip}\nlogging trap {level}\nlogging on", "notes": "Syslog远程日志"},
    {"vendor": "cisco", "task": "VTP配置", "commands": "vtp domain {domain}\nvtp mode {server|client|transparent}\nvtp password {password}", "notes": "VLAN中继协议"},
    {"vendor": "cisco", "task": "端口镜像", "commands": "monitor session 1 source interface {source_interface}\nmonitor session 1 destination interface {dest_interface}", "notes": "SPAN端口镜像配置"},

    {"vendor": "huawei", "task": "端口状态检查", "commands": "display interface brief\ndisplay ip interface brief", "notes": "华为端口状态"},
    {"vendor": "huawei", "task": "VLAN配置", "commands": "vlan {vlan_id}\ndescription {vlan_name}\ninterface {interface}\nport link-type access\nport default vlan {vlan_id}", "notes": "华为VLAN配置"},
    {"vendor": "huawei", "task": "Trunk配置", "commands": "interface {interface}\nport link-type trunk\nport trunk allow-pass vlan {vlan_list}", "notes": "华为Trunk配置"},
    {"vendor": "huawei", "task": "路由表检查", "commands": "display ip routing-table\ndisplay ospf peer", "notes": "华为路由表"},
    {"vendor": "huawei", "task": "链路聚合", "commands": "interface eth-trunk {num}\nmode lacp\ninterface {interface}\neth-trunk {num}", "notes": "华为LACP"},
    {"vendor": "huawei", "task": "DHCP配置", "commands": "dhcp enable\nip pool {name}\nnetwork {network} mask {mask}\ngateway-list {gateway}\ninterface {interface}\ndhcp select global", "notes": "华为DHCP服务器"},
    {"vendor": "huawei", "task": "OSPF配置", "commands": "ospf {pid}\narea {area}\nnetwork {network} {wildcard}\ninterface {interface}\nospf enable {pid} area {area}", "notes": "华为OSPF"},
    {"vendor": "huawei", "task": "BGP配置", "commands": "bgp {asn}\npeer {ip} as-number {remote_asn}\nnetwork {network} {mask}", "notes": "华为BGP"},
    {"vendor": "huawei", "task": "静态路由", "commands": "ip route-static {destination} {mask} {next_hop}", "notes": "华为静态路由"},
    {"vendor": "huawei", "task": "SSH配置", "commands": "stelnet server enable\nssh user {user}\nssh user {user} authentication-type password\nssh user {user} service-type stelnet", "notes": "华为SSH服务"},
    {"vendor": "huawei", "task": "SNMP配置", "commands": "snmp-agent community read {community}\nsnmp-agent sys-info location {location}\nsnmp-agent sys-info contact {email}", "notes": "华为SNMP"},

    {"vendor": "h3c", "task": "端口状态检查", "commands": "display interface brief\ndisplay ip interface brief", "notes": "H3C端口"},
    {"vendor": "h3c", "task": "VLAN配置", "commands": "vlan {vlan_id}\nname {vlan_name}\ninterface {interface}\nport access vlan {vlan_id}", "notes": "H3C VLAN"},
    {"vendor": "h3c", "task": "Trunk配置", "commands": "interface {interface}\nport link-type trunk\nport trunk permit vlan {vlan_list}", "notes": "H3C Trunk"},
    {"vendor": "h3c", "task": "SSH配置", "commands": "ssh server enable\nlocal-user {user} class manage\npassword simple {password}\nservice-type ssh\nauthorization-attribute user-role network-admin", "notes": "H3C SSH"},
    {"vendor": "h3c", "task": "OSPF配置", "commands": "ospf {pid}\narea {area}\nnetwork {network} {wildcard}", "notes": "H3C OSPF"},
    {"vendor": "h3c", "task": "链路聚合", "commands": "interface bridge-aggregation {num}\ninterface {interface}\nport link-aggregation group {num}", "notes": "H3C链路聚合"},
    {"vendor": "h3c", "task": "端口镜像", "commands": "mirroring-group 1 local\nmirroring-group 1 mirroring-port {interface} both\nmirroring-group 1 monitor-port {interface}", "notes": "H3C端口镜像"},
    {"vendor": "h3c", "task": "静态路由", "commands": "ip route-static {destination} {mask} {next_hop}", "notes": "H3C静态路由"},

    {"vendor": "juniper", "task": "端口状态检查", "commands": "show interfaces terse\nshow interfaces descriptions", "notes": "Juniper端口"},
    {"vendor": "juniper", "task": "VLAN配置", "commands": "set vlans {name} vlan-id {id}\nset interfaces {interface} unit 0 family ethernet-switching vlan members {name}", "notes": "Juniper VLAN"},
    {"vendor": "juniper", "task": "OSPF配置", "commands": "set protocols ospf area 0 interface {interface}\nset protocols ospf area {area} interface {interface} passive", "notes": "Juniper OSPF"},
    {"vendor": "juniper", "task": "BGP配置", "commands": "set protocols bgp group {group} type external\nset protocols bgp group {group} peer-as {asn}\nset protocols bgp group {group} neighbor {ip}", "notes": "Juniper BGP"},
    {"vendor": "juniper", "task": "防火墙过滤", "commands": "set firewall family inet filter {name} term {term} from source-address {src}\nset firewall family inet filter {name} term {term} then accept", "notes": "Juniper防火墙过滤"},
    {"vendor": "juniper", "task": "SNMP配置", "commands": "set snmp community {community}\nset snmp location {location}\nset snmp contact {email}", "notes": "Juniper SNMP"},

    {"vendor": "fortinet", "task": "端口状态检查", "commands": "get system interface physical\nget system interface", "notes": "FortiGate端口"},
    {"vendor": "fortinet", "task": "VLAN配置", "commands": "config system interface\nedit {vlan_name}\nset vdom root\nset vlanid {vlan_id}\nset interface {parent_interface}\nend", "notes": "FortiGate VLAN"},
    {"vendor": "fortinet", "task": "防火墙策略", "commands": "config firewall policy\nedit {id}\nset srcintf {src}\nset dstintf {dst}\nset srcaddr {src_addr}\nset dstaddr {dst_addr}\nset action accept\nset service ALL\nend", "notes": "FortiGate防火墙策略"},
    {"vendor": "fortinet", "task": "静态路由", "commands": "config router static\nedit 1\nset device {interface}\nset gateway {gateway}\nset dst {destination} {mask}\nend", "notes": "FortiGate静态路由"},
    {"vendor": "fortinet", "task": "IPSec VPN", "commands": "config vpn ipsec phase1-interface\nedit {name}\nset interface {wan_interface}\nset peertype any\nset net-device enable\nset proposal aes256-sha256\nset dhgroup 14\nset remote-gw {peer_ip}\nset psksecret {psk}\nnext\nend", "notes": "FortiGate IPSec VPN配置"},
    {"vendor": "fortinet", "task": "SSH配置", "commands": "config system global\nset admin-ssh-port 22\nset admin-ssh-password-auth enable\nend", "notes": "FortiGate SSH管理"},
    {"vendor": "fortinet", "task": "SNMP配置", "commands": "config system snmp community\nedit 1\nset name {community}\nset events v1 v2c\nconfig hosts\nedit 1\nset ip {nms_ip}\nnext\nend\nnext\nend", "notes": "FortiGate SNMP"},

    {"vendor": "ruijie", "task": "端口状态检查", "commands": "show interface status\nshow interface", "notes": "锐捷端口"},
    {"vendor": "ruijie", "task": "VLAN配置", "commands": "vlan {vlan_id}\nname {vlan_name}\ninterface {interface}\nswitchport access vlan {vlan_id}", "notes": "锐捷VLAN"},
    {"vendor": "ruijie", "task": "Trunk配置", "commands": "interface {interface}\nswitchport mode trunk\nswitchport trunk allowed vlan {vlan_list}", "notes": "锐捷Trunk"},
    {"vendor": "ruijie", "task": "链路聚合", "commands": "interface aggregateport {num}\ninterface {interface}\nport aggregateport {num}", "notes": "锐捷链路聚合"},
    {"vendor": "ruijie", "task": "SSH配置", "commands": "enable service ssh-server\nip ssh version 2\nline vty 0 4\ntransport input ssh", "notes": "锐捷SSH"},

    {"vendor": "hpe", "task": "端口状态检查", "commands": "show interfaces brief\nshow interface status", "notes": "HPE/Aruba端口"},
    {"vendor": "hpe", "task": "VLAN配置", "commands": "vlan {vlan_id}\nname {vlan_name}\ninterface {interface}\nuntagged vlan {vlan_id}", "notes": "HPE/Aruba VLAN（ProVision）"},
    {"vendor": "hpe", "task": "Trunk配置", "commands": "interface {interface}\ntagged vlan {vlan_list}", "notes": "HPE/Aruba Trunk"},

    {"vendor": "*", "task": "系统信息", "commands": "show version\nshow running-config | include hostname", "notes": "查看设备基本信息"},
    {"vendor": "*", "task": "配置备份", "commands": "copy running-config startup-config\nshow running-config", "notes": "备份当前配置"},
    {"vendor": "*", "task": "日志检查", "commands": "show logging\nshow log", "notes": "检查系统日志"},
    {"vendor": "*", "task": "CPU检查", "commands": "show process cpu\nshow cpu usage", "notes": "查看CPU使用率"},
    {"vendor": "*", "task": "内存检查", "commands": "show memory\nshow process memory", "notes": "查看内存使用率"},
    {"vendor": "*", "task": "接口错误检查", "commands": "show interface counters errors\nshow interface statistics", "notes": "检查接口错误计数"},
    {"vendor": "*", "task": "ARP表", "commands": "show arp\nshow ip arp", "notes": "查看ARP表"},
    {"vendor": "*", "task": "MAC地址表", "commands": "show mac address-table\nshow mac-address-table", "notes": "查看MAC地址表"},
    {"vendor": "*", "task": "环境状态", "commands": "show environment\nshow env all", "notes": "检查设备温度/电源/风扇"},
    {"vendor": "*", "task": "LLDP邻居", "commands": "show lldp neighbors detail\nshow lldp neighbors", "notes": "查看LLDP邻居信息"},
    {"vendor": "*", "task": "CDP邻居", "commands": "show cdp neighbors detail\nshow cdp neighbors", "notes": "查看CDP邻居信息"},
    {"vendor": "*", "task": "启动配置", "commands": "show startup-config", "notes": "查看启动配置"},
    {"vendor": "*", "task": "运行配置", "commands": "show running-config", "notes": "查看当前运行配置"},
    {"vendor": "*", "task": "配置文件比较", "commands": "show archive config differences\nshow diff", "notes": "比较配置差异"},
    {"vendor": "*", "task": "License查看", "commands": "show license\nshow license summary", "notes": "查看许可证信息"},
    {"vendor": "*", "task": "光模块信息", "commands": "show interface transceiver\nshow interface optics", "notes": "查看光模块状态"},
    {"vendor": "*", "task": "PoE信息", "commands": "show power inline\nshow poe", "notes": "查看PoE供电状态"},
]


def init_knowledge_base(force: bool = False) -> int:
    """初始化知识库，首次自动加载预设模板"""
    conn = _get_db()
    if force:
        conn.execute("DELETE FROM configs")
    count = conn.execute("SELECT COUNT(*) FROM configs").fetchone()[0]
    if count == 0:
        return batch_add(_PRESET_CONFIGS)
    return count


def add_config(vendor: str, task: str, commands: str, notes: str = "") -> dict:
    """添加单条配置"""
    conn = _get_db()
    conn.execute(
        "INSERT INTO configs (vendor, task, commands, notes) VALUES (?, ?, ?, ?)",
        (vendor, task, commands, notes)
    )
    conn.commit()
    return {"added": 1, "total": conn.execute("SELECT COUNT(*) FROM configs").fetchone()[0]}


def batch_add(entries: list[dict]) -> int:
    """批量添加配置"""
    if not entries:
        return 0
    conn = _get_db()
    for e in entries:
        conn.execute(
            "INSERT INTO configs (vendor, task, commands, notes) VALUES (?, ?, ?, ?)",
            (e.get("vendor", "*"), e.get("task", ""), e.get("commands", ""), e.get("notes", ""))
        )
    conn.commit()
    return len(entries)


def search_commands(query: str, vendor: str = "*", top_k: int = 10) -> list[dict]:
    """FTS5 全文搜索"""
    if not query or not query.strip():
        return []
    conn = _get_db()
    # 清洗查询词，构造 FTS5 查询语法
    terms = query.strip().split()
    fts_query = " AND ".join(f'"{t}"' for t in terms if t)
    if not fts_query:
        return []

    if vendor == "*":
        rows = conn.execute(
            "SELECT c.vendor, c.task, c.commands, c.notes FROM configs_fts f "
            "JOIN configs c ON f.rowid = c.id "
            "WHERE configs_fts MATCH ? ORDER BY rank LIMIT ?",
            (fts_query, top_k)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT c.vendor, c.task, c.commands, c.notes FROM configs_fts f "
            "JOIN configs c ON f.rowid = c.id "
            "WHERE configs_fts MATCH ? AND (c.vendor = ? OR c.vendor = '*') "
            "ORDER BY rank LIMIT ?",
            (fts_query, vendor, top_k)
        ).fetchall()

    return [dict(r) for r in rows]


def get_all_configs() -> list[dict]:
    conn = _get_db()
    rows = conn.execute("SELECT vendor, task, commands, notes FROM configs ORDER BY vendor, task").fetchall()
    return [dict(r) for r in rows]


def knowledge_stats() -> dict:
    conn = _get_db()
    total = conn.execute("SELECT COUNT(*) FROM configs").fetchone()[0]
    rows = conn.execute("SELECT vendor, COUNT(*) as cnt FROM configs GROUP BY vendor").fetchall()
    vendors = {r["vendor"]: r["cnt"] for r in rows}
    return {"total": total, "vendors": vendors, "backend": "SQLite FTS5"}


def import_from_xlsx(filepath_or_bytes) -> dict:
    """从 CSV 或 XLSX 文件导入"""
    entries = _parse_import(filepath_or_bytes)
    if entries:
        added = batch_add(entries)
        total = _get_db().execute("SELECT COUNT(*) FROM configs").fetchone()[0]
        return {"added": added, "total": total}
    return {"added": 0, "error": "无有效数据"}


def _parse_import(source) -> list[dict]:
    entries = []
    if isinstance(source, bytes):
        text = source.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            entry = _extract_row(row)
            if entry:
                entries.append(entry)
    else:
        path = str(source)
        ext = os.path.splitext(path.lower())[1]
        if ext == '.csv':
            with open(path, 'r', encoding='utf-8-sig') as f:
                for row in csv.DictReader(f):
                    entry = _extract_row(row)
                    if entry:
                        entries.append(entry)
        else:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path)
                ws = wb.active
                headers = [str(cell.value or '').strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                vi = next((i for i, h in enumerate(headers) if h.lower() in ('vendor', '厂商')), 0)
                ti = next((i for i, h in enumerate(headers) if h.lower() in ('task', '任务')), 1)
                ci = next((i for i, h in enumerate(headers) if h.lower() in ('commands', '命令')), 2)
                ni = next((i for i, h in enumerate(headers) if h.lower() in ('notes', '备注')), 3)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    v = str(row[vi] or '').strip().lower() if vi < len(row) else ''
                    t = str(row[ti] or '').strip() if ti < len(row) else ''
                    c = str(row[ci] or '').strip() if ci < len(row) else ''
                    n = str(row[ni] or '').strip() if ni < len(row) else ''
                    if v and t and c:
                        entries.append({"vendor": v, "task": t, "commands": c, "notes": n})
            except ImportError:
                pass
    return entries


def _extract_row(row: dict) -> dict | None:
    v = (row.get('vendor') or row.get('厂商') or '').strip().lower()
    t = (row.get('task') or row.get('任务') or '').strip()
    c = (row.get('commands') or row.get('命令') or '').strip()
    n = (row.get('notes') or row.get('备注') or '').strip()
    return {"vendor": v, "task": t, "commands": c, "notes": n} if v and t and c else None


def delete_all():
    conn = _get_db()
    conn.execute("DELETE FROM configs")
    conn.commit()


def knowledge_summary() -> dict:
    return knowledge_stats()


# 向后兼容
search_configs = search_commands
add_command = add_config
